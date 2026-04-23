import os
import shutil
from datetime import datetime
import openpyxl
from database import get_db, get_distinct_planilhas

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PLANILHAS_DIR = os.path.join(BASE_DIR, 'planilhas_excel')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')

VMB_COL = 10  # Coluna "VALOR DE MERCADO (VMB)"


def export_all():
    """
    Para cada planilha com bens avaliados, cria uma cópia do Excel original
    na pasta output/ com a coluna VMB preenchida.
    Retorna lista dos arquivos gerados.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    date_str = datetime.now().strftime('%Y%m%d_%H%M')
    exported = []

    conn = get_db()
    for planilha in get_distinct_planilhas():
        rows = conn.execute("""
            SELECT a.row_index, r.valor_mercado, r.metodologia
            FROM assets a
            JOIN reviews r ON a.id = r.asset_id
            WHERE a.planilha = ?
        """, (planilha,)).fetchall()

        if not rows:
            continue

        src = os.path.join(PLANILHAS_DIR, planilha + '.xlsx')
        if not os.path.exists(src):
            continue

        dst = os.path.join(OUTPUT_DIR, f"{planilha}_avaliado_{date_str}.xlsx")
        shutil.copy2(src, dst)

        wb = openpyxl.load_workbook(dst)
        ws = wb.active
        for row_index, valor_mercado, metodologia in rows:
            ws.cell(row=row_index, column=VMB_COL).value = valor_mercado
            ws.cell(row=row_index, column=VMB_COL + 1).value = metodologia or 'M1'
        wb.save(dst)
        wb.close()
        exported.append(os.path.basename(dst))

    conn.close()
    return exported
