import os
import openpyxl
from database import get_db

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PLANILHAS_DIR = os.path.join(BASE_DIR, 'planilhas_excel')

# Colunas (índice 1-based conforme planilha)
COL_NATUREZA   = 1
COL_MATERIAL   = 2
COL_NRP        = 3
COL_TIPO       = 4
COL_MARCA      = 5
COL_MODELO     = 6
COL_DATA_TOMB  = 7
COL_VAL_CONT   = 8
COL_VAL_ATUAL  = 9


def _find_data_start(ws):
    """Detecta a primeira linha de dados buscando a linha de cabeçalho que contém 'NRP'."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        for cell in row:
            if cell is not None and 'NRP' in str(cell).upper():
                return i + 1  # dados começam na linha seguinte ao cabeçalho
    return 8  # fallback conservador


def _rows_from_workbook(path):
    """Abre o xlsx e retorna lista de tuplas (row_index, values_tuple)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    data_start = _find_data_start(ws)
    result = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=data_start, values_only=True),
        start=data_start,
    ):
        nrp = _normalize(row[COL_NRP - 1])
        if not nrp:
            continue
        result.append((row_idx, row))
    wb.close()
    return result


def _normalize(v):
    """Converte valor de célula para string ou None."""
    if v is None:
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%d/%m/%Y')
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(v)
    return str(v).strip() or None


def _to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _count_for_planilha(conn, planilha):
    return conn.execute(
        "SELECT COUNT(*) FROM assets WHERE planilha = ?", (planilha,)
    ).fetchone()[0]


def load_excel_files():
    """
    Importa todas as planilhas para o banco.
    Idempotente por planilha: planilhas já importadas são ignoradas.
    Usa INSERT OR IGNORE para tolerar re-execuções parciais.
    """
    total = 0
    for fname in sorted(os.listdir(PLANILHAS_DIR)):
        if not fname.lower().endswith('.xlsx'):
            continue
        path = os.path.join(PLANILHAS_DIR, fname)
        planilha = os.path.splitext(fname)[0]

        conn = get_db()
        if _count_for_planilha(conn, planilha) > 0:
            conn.close()
            continue  # planilha já importada

        # Lê o arquivo Excel (sem conexão de banco aberta)
        conn.close()
        try:
            rows = _rows_from_workbook(path)
        except Exception as e:
            print(f"Erro ao ler {fname}: {e}")
            continue

        if not rows:
            print(f"  {fname}: sem dados")
            continue

        conn = get_db()
        count_file = 0
        for row_idx, row in rows:
            conn.execute("""
                INSERT OR IGNORE INTO assets
                    (planilha, row_index, natureza_despesa, material, nrp, tipo,
                     marca, modelo, data_tombamento, valor_contabil, valor_atual)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                planilha,
                row_idx,
                _normalize(row[COL_NATUREZA - 1]),
                _normalize(row[COL_MATERIAL - 1]),
                _normalize(row[COL_NRP - 1]),
                _normalize(row[COL_TIPO - 1]),
                _normalize(row[COL_MARCA - 1]),
                _normalize(row[COL_MODELO - 1]),
                _normalize(row[COL_DATA_TOMB - 1]),
                _to_float(row[COL_VAL_CONT - 1]),
                _to_float(row[COL_VAL_ATUAL - 1]),
            ))
            count_file += 1
        conn.commit()
        conn.close()
        total += count_file
        print(f"  {fname}: {count_file} bens")

    if total:
        print(f"Importação concluída: {total} bens carregados.")
    return total
